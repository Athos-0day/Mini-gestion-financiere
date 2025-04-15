from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
import os
import shutil
import pandas as pd
from datetime import datetime

# Dossier des données
DOSSIER_DATA = "data"

# Fichier flag pour indiquer que l'initialisation a été faite
FLAG_FILE = os.path.join(DOSSIER_DATA, "first_use.flag")

# Structure des fichiers et des colonnes
STRUCTURES = {
    "factures.xlsx": ["ID", "Date", "Client", "Description", "Type", "Statut", "Email", "Montant HT (€)", "Montant TTC (€)"],
    "depenses.xlsx": ["ID", "Date", "Nom", "Description", "Catégorie", "Email", "Montant HT (€)", "Montant TTC (€)"],
    "historique.xlsx": ["ID", "Date", "Nom", "Description", "Type", "Email", "Montant HT (€)", "Montant TTC (€)"]
}

# Taux de TVA (ex: 20%)
TAUX_TVA = 0.20

# Fonction de sauvegarde
def sauvegarder_fichier(fichier):
    now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    nom_sans_ext = os.path.splitext(fichier)[0]
    nom_sauvegarde = f"{nom_sans_ext}_{now}.xlsx"
    dossier_sauvegarde = os.path.join(DOSSIER_DATA, "sauvegardes")
    os.makedirs(dossier_sauvegarde, exist_ok=True)
    chemin_original = os.path.join(DOSSIER_DATA, fichier)
    chemin_sauvegarde = os.path.join(dossier_sauvegarde, nom_sauvegarde)
    shutil.copy2(chemin_original, chemin_sauvegarde)
    print(f"   Sauvegarde de {fichier} effectuée dans 'sauvegardes/{nom_sauvegarde}'")

# Fonction pour appliquer des bordures et des styles
def appliquer_styles(feuille):
    # Définir le style de bordure
    fine = Side(border_style="thin", color="000000")  # bordure fine (Noir)
    bordure = Border(top=fine, left=fine, right=fine, bottom=fine)

    # Style de police (gras pour les en-têtes et Montant TTC)
    en_tete_font = Font(bold=True)
    montant_ttc_font = Font(bold=True, color="FF0000")  # en gras et rouge pour le Montant TTC
    
    # Style de fond pour les en-têtes (couleur grise)
    en_tete_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Appliquer les styles sur les en-têtes (première ligne)
    for cell in feuille[1]:
        cell.font = en_tete_font
        cell.fill = en_tete_fill
        cell.border = bordure

    # Appliquer les bordures à toutes les cellules et en gras les Montants TTC
    for row in feuille.iter_rows(min_row=2, max_col=len(feuille[1]), max_row=feuille.max_row):
        for cell in row:
            # Appliquer la bordure
            cell.border = bordure
            
            # Mettre en gras le montant TTC (ici on suppose que c'est la colonne 6 : "Montant TTC (€)")
            if cell.column == 6:  # La colonne "Montant TTC (€)" est la 6ème colonne
                cell.font = montant_ttc_font

    # Élargir les colonnes en fonction du contenu
    for col in range(1, feuille.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in feuille.iter_rows(min_row=1, max_row=feuille.max_row, min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        feuille.column_dimensions[column_letter].width = adjusted_width

# Fonction d'initialisation des fichiers Excel avec style
def initialiser_excel(annee=None):
    # Vérifie si c'est la première utilisation
    if os.path.exists(FLAG_FILE):
        print("Les fichiers ont déjà été initialisés.\n")
        return

    if annee is None:
        annee = int(input("Veuillez entrer l'année de départ (ex: 2024) : "))

    mois = ["Septembre", "Octobre", "Novembre", "Décembre", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet"]
    mois_num = [9, 10, 11, 12, 1, 2, 3, 4, 5, 6, 7]

    print("\n--- Initialisation des fichiers Excel ---\n")

    for fichier, colonnes in STRUCTURES.items():
        chemin = os.path.join(DOSSIER_DATA, fichier)

        if os.path.exists(chemin):
            print(f"Le fichier {fichier} existe déjà. Ignorer l'initialisation.")
            continue
        
        print(f"Création du fichier : {fichier}")
        try:
            with pd.ExcelWriter(chemin, engine='openpyxl') as writer:
                for i, m in enumerate(mois):
                    annee_mois = annee if mois_num[i] >= 9 else annee + 1
                    feuille_mois = f"{m} {annee_mois}"
                    print(f"   Création de la feuille : {feuille_mois}")
                    
                    # Créer un DataFrame vide avec les colonnes définies
                    df = pd.DataFrame(columns=colonnes)
                    df.to_excel(writer, sheet_name=feuille_mois, index=False)
                    
                    # Appliquer les styles après l'écriture du DataFrame
                    feuille = writer.sheets[feuille_mois]
                    appliquer_styles(feuille)

            print(f"   Fichier {fichier} créé avec succès.\n")
        except Exception as e:
            print(f"   Erreur lors de la création du fichier {fichier} : {e}\n")

    # Marquer la fin de l'initialisation en créant le fichier flag
    with open(FLAG_FILE, "w") as f:
        f.write("Fichiers Excel initialisés.\n")
    
    print("--- Initialisation terminée ---\n")

# Fonction pour forcer l'initialisation (en cas de réinitialisation manuelle)
def reinitialiser_excel(annee=None):
    # Supprime le fichier flag pour forcer l'initialisation
    if os.path.exists(FLAG_FILE):
        os.remove(FLAG_FILE)
        print("Réinitialisation forcée. Le fichier flag a été supprimé.\n")
        initialiser_excel(annee)
    else:
        print("L'initialisation n'a pas encore eu lieu. Création des fichiers Excel.\n")
        initialiser_excel(annee)

