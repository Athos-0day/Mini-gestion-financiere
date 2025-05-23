import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment, NamedStyle, numbers, PatternFill
import os
import datetime
from utils.validation import email_valide, montant_valide, statut_valide, categorie_valide
from openpyxl.utils import get_column_letter

DOSSIER_DATA = "data"
FICHIER_FACTURES = "factures.xlsx"
FICHIER_HISTORIQUE = "historique.xlsx"
FICHIER_DEPENSES = "depenses.xlsx"
TAUX_TVA = 0.20

def ajouter_facture(facture):
    try:
        date_facture = pd.to_datetime(facture["Date"], format="%Y-%m-%d")
    except Exception:
        raise ValueError(f"Date invalide : {facture['Date']}")

    if not email_valide(facture.get("Email")):
        raise ValueError(f"Adresse e-mail invalide : {facture.get('Email')}")

    if not montant_valide(facture.get("Montant TTC (€)")):
        raise ValueError(f"Montant TTC invalide : {facture.get('Montant TTC (€)')}")

    if not statut_valide(facture.get("Statut")):
        raise ValueError(f"Statut invalide : {facture.get('Statut')} (attendu : Envoyée, Payée, En attente, Annulée)")

    try:
        montant_ttc_val = float(facture["Montant TTC (€)"])
        montant_ht_val = round(montant_ttc_val / (1 + TAUX_TVA), 2)

        facture["Montant TTC (€)"] = f"{montant_ttc_val:.2f} €"
        facture["Montant HT (€)"] = f"{montant_ht_val:.2f} €"
    except Exception:
        raise ValueError(f"Erreur lors du calcul du montant HT à partir du montant TTC : {facture.get('Montant TTC (€)')}")

    try:
        mois = date_facture.month
        annee = date_facture.year

        mois_en = date_facture.strftime("%B")
        mois_fr_dict = {
            "September": "Septembre", "October": "Octobre", "November": "Novembre",
            "December": "Décembre", "January": "Janvier", "February": "Février",
            "March": "Mars", "April": "Avril", "May": "Mai", "June": "Juin", "July": "Juillet"
        }

        mois_fr = mois_fr_dict[mois_en]
        nom_feuille = f"{mois_fr} {annee}"
        chemin_fichier = os.path.join(DOSSIER_DATA, FICHIER_FACTURES)

        wb = load_workbook(chemin_fichier)
        if nom_feuille not in wb.sheetnames:
            raise ValueError(f"La feuille '{nom_feuille}' n'existe pas.")
        
        ws = wb[nom_feuille]

        # Génération ID
        lignes = list(ws.iter_rows(min_row=2, values_only=True))
        prefix = f"F{annee}{mois:02d}"
        existing_ids = [str(row[0]) for row in lignes if row[0] and str(row[0]).startswith(prefix)]
        prochain_num = max([int(id_[7:]) for id_ in existing_ids], default=0) + 1
        facture_id = f"{prefix}{prochain_num}"
        facture["ID"] = facture_id

        # Récupérer ordre des colonnes
        colonnes = [cell.value for cell in ws[1]]

        # Créer la ligne à ajouter
        nouvelle_ligne = [facture.get(col, "") for col in colonnes]
        next_row = ws.max_row + 1
        ws.append(nouvelle_ligne)

        # === STYLE APPLIQUÉ À LA LIGNE ===
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        align = Alignment(horizontal="center", vertical="center")
        bold = Font(bold=True)

        for col_idx, value in enumerate(nouvelle_ligne, start=1):
            cell = ws.cell(row=next_row, column=col_idx)
            cell.alignment = align
            cell.border = border

            # Gras pour TTC uniquement
            if colonnes[col_idx - 1] == "Montant TTC (€)":
                cell.font = bold

        wb.save(chemin_fichier)

        print("\nFacture ajoutée avec succès :")
        print("┌──────────────────────────────────────────")
        print(f"│ ID       : {facture['ID']}")
        print(f"│ Date     : {facture['Date']}")
        print(f"│ Client   : {facture['Client']}")
        print(f"│ Montant  : {facture['Montant TTC (€)']} / {facture['Montant HT (€)']}")
        print(f"│ Feuille  : {nom_feuille}")
        print("└──────────────────────────────────────────\n")

        # Ajout à l'historique
        facture["Nom"] = facture["Client"]  # pour cohérence avec le champ "Nom" dans historique
        facture["Type"] = "Facture"
        ajouter_historique(facture, type_="Facture")

    except Exception as e:
        raise RuntimeError(f"Erreur lors de l'ajout de la facture : {e}")

# Fonction pour ajouter une dépense
def ajouter_depense(depense):
    try:
        date_depense = pd.to_datetime(depense["Date"], format="%Y-%m-%d")
    except Exception:
        raise ValueError(f"Date invalide : {depense['Date']}")

    if not email_valide(depense.get("Email")):
        raise ValueError(f"Adresse e-mail invalide : {depense.get('Email')}")

    if not montant_valide(depense.get("Montant TTC (€)")):
        raise ValueError(f"Montant TTC invalide : {depense.get('Montant TTC (€)')}")

    if not categorie_valide(depense.get("Catégorie")):
        raise ValueError(f"Catégorie invalide : {depense.get('Catégorie')}")

    try:
        montant_ttc_val = float(depense["Montant TTC (€)"])
        montant_ht_val = round(montant_ttc_val / (1 + TAUX_TVA), 2)

        depense["Montant TTC (€)"] = f"{montant_ttc_val:.2f} €"
        depense["Montant HT (€)"] = f"{montant_ht_val:.2f} €"
    except Exception:
        raise ValueError(f"Erreur lors du calcul du montant HT à partir du montant TTC : {depense.get('Montant TTC (€)')}")

    try:
        mois = date_depense.month
        annee = date_depense.year

        mois_en = date_depense.strftime("%B")
        mois_fr_dict = {
            "September": "Septembre", "October": "Octobre", "November": "Novembre",
            "December": "Décembre", "January": "Janvier", "February": "Février",
            "March": "Mars", "April": "Avril", "May": "Mai", "June": "Juin", "July": "Juillet"
        }

        mois_fr = mois_fr_dict[mois_en]
        nom_feuille = f"{mois_fr} {annee}"
        chemin_fichier = os.path.join(DOSSIER_DATA, FICHIER_DEPENSES)

        wb = load_workbook(chemin_fichier)
        if nom_feuille not in wb.sheetnames:
            raise ValueError(f"La feuille '{nom_feuille}' n'existe pas.")
        
        ws = wb[nom_feuille]

        # Génération ID
        lignes = list(ws.iter_rows(min_row=2, values_only=True))
        prefix = f"D{annee}{mois:02d}"
        existing_ids = [str(row[0]) for row in lignes if row[0] and str(row[0]).startswith(prefix)]
        prochain_num = max([int(id_[7:]) for id_ in existing_ids], default=0) + 1
        depense_id = f"{prefix}{prochain_num}"
        depense["ID"] = depense_id

        # Récupérer ordre des colonnes
        colonnes = [cell.value for cell in ws[1]]

        # Créer la ligne à ajouter
        nouvelle_ligne = [depense.get(col, "") for col in colonnes]
        next_row = ws.max_row + 1
        ws.append(nouvelle_ligne)

        # === STYLE APPLIQUÉ À LA LIGNE ===
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        align = Alignment(horizontal="center", vertical="center")
        bold = Font(bold=True)

        for col_idx, value in enumerate(nouvelle_ligne, start=1):
            cell = ws.cell(row=next_row, column=col_idx)
            cell.alignment = align
            cell.border = border

            # Gras pour TTC uniquement
            if colonnes[col_idx - 1] == "Montant TTC (€)":
                cell.font = bold

        wb.save(chemin_fichier)

        print("\nDépense ajoutée avec succès :")
        print("┌──────────────────────────────────────────")
        print(f"│ ID       : {depense['ID']}")
        print(f"│ Date     : {depense['Date']}")
        print(f"│ Fournisseur : {depense['Nom']}")
        print(f"│ Montant  : {depense['Montant TTC (€)']} € TTC / {depense['Montant HT (€)']} € HT")
        print(f"│ Feuille  : {nom_feuille}")
        print("└──────────────────────────────────────────\n")

        # Ajout à l'historique
        depense["Nom"] = depense["Nom"]  # déjà le bon champ
        depense["Type"] = "Dépense"
        ajouter_historique(depense, type_="Dépense")

    except Exception as e:
        raise RuntimeError(f"Erreur lors de l'ajout de la dépense : {e}")

# Dictionnaire de traduction des mois de l'anglais vers le français
mois_fr = {
    "January": "Janvier",
    "February": "Février",
    "March": "Mars",
    "April": "Avril",
    "May": "Mai",
    "June": "Juin",
    "July": "Juillet",
    "August": "Août",
    "September": "Septembre",
    "October": "Octobre",
    "November": "Novembre",
    "December": "Décembre"
}

def ajouter_historique(element, type_):
    try:
        chemin_fichier = os.path.join(DOSSIER_DATA, "historique.xlsx")

        # Crée le fichier s'il n'existe pas
        if not os.path.exists(chemin_fichier):
            df_init = pd.DataFrame(columns=["ID", "DateAjout", "DateReel", "Heure", "Nom", "Type", "Montant (€)"])
            df_init.to_excel(chemin_fichier, index=False)

        # Nettoyage du montant (supprime € et espace)
        montant_str = element["Montant TTC (€)"]
        montant_float = float(montant_str.replace("€", "").replace(" ", "").strip())

        # Inverser le montant si c'est une dépense
        if type_ == "Dépense":
            montant_float = -abs(montant_float)
        else:
            montant_float = abs(montant_float)

        # Déterminer le mois et l'année à partir de DateAjout
        date_ajout = datetime.datetime.strptime(element["Date"], "%Y-%m-%d")
        mois_anglais = date_ajout.strftime("%B")  # Mois en anglais
        mois = mois_fr[mois_anglais]  # Traduction en français
        annee = date_ajout.year

        # Créer un nom de feuille basé sur le mois et l'année
        nom_feuille = f"{mois} {annee}"

        # Charger le fichier Excel
        wb = load_workbook(chemin_fichier)

        # Créer la feuille si elle n'existe pas
        if nom_feuille not in wb.sheetnames:
            ws = wb.create_sheet(nom_feuille)
            ws.append(["ID", "DateAjout", "DateReel", "Heure", "Nom","Type", "Montant (€)"])  # Ajouter les en-têtes
        else:
            ws = wb[nom_feuille]

        # Ajouter les nouvelles données à la feuille
        nouvelle_ligne = [
            element["ID"],
            element["Date"],  # Date de la facture ou dépense
            datetime.datetime.now().strftime("%Y-%m-%d"),
            datetime.datetime.now().strftime("%H:%M:%S"),
            element["Nom"],
            type_,
            montant_float
        ]
        ws.append(nouvelle_ligne)

        # Appliquer le formatage sur la nouvelle ligne ajoutée
        last_row = ws.max_row

        # Centrer toutes les cellules de la ligne ajoutée
        for cell in ws[last_row]:
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Mettre le montant en euros et en gras pour le montant TTC
        montant_cell = ws.cell(row=last_row, column=7)  # Le montant est dans la 7ème colonne
        
        # Format personnalisé Euro, exemple : 1 000,00 € ou 1 000.00 €
        montant_cell.number_format = '#,##0.00 €'  # Ajout du format personnalisé pour l'Euro
        montant_cell.font = Font(bold=True)  # Gras pour le montant TTC

        # Appliquer la couleur du texte en fonction du type (rouge pour Dépense, vert pour Facture)
        if type_ == "Dépense":
            montant_cell.font = Font(color="FF0000", bold=True)  # Rouge pour les dépenses
        else:
            montant_cell.font = Font(color="00FF00", bold=True)  # Vert pour les factures

        # Sauvegarder les changements dans le fichier Excel
        wb.save(chemin_fichier)

    except Exception as e:
        raise RuntimeError(f"Erreur lors de l'ajout à l'historique : {e}")